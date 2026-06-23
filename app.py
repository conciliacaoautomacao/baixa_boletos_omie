import streamlit as st
import pandas as pd
import fitz
import re
import io
import os
from datetime import date, datetime, timedelta
from supabase import create_client
from openpyxl import load_workbook
from copy import copy
import time
from zoneinfo import ZoneInfo
import unicodedata

# =============================
# CONFIG
# =============================
st.set_page_config(
    page_title="Baixa Boletos Omie",
    page_icon="📄",
    layout="wide"
)

SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

MODELO_EXCEL = "modelo/Omie_Contas_Pagar_v1_1_5.xlsx"


# =============================
# FUNÇÕES
# =============================

def data_hoje_br_date():
    return datetime.now(ZoneInfo("America/Sao_Paulo")).date()

def buscar_config_padrao(tipo_operacao):
    res = (
        supabase.table("config_padrao_omie")
        .select("*")
        .eq("tipo_operacao", tipo_operacao)
        .eq("ativo", True)
        .limit(1)
        .execute()
    )

    if not res.data:
        raise Exception(f"Nenhuma configuração ativa encontrada para {tipo_operacao}")

    return res.data[0]
    
def gerar_remessa_id():
    try:
        res = (
            supabase.table("boletos_extraidos")
            .select("remessa_id")
            .not_.is_("remessa_id", "null")
            .execute()
        )

        df = pd.DataFrame(res.data)

        numeros = []

        if not df.empty:
            for remessa in df["remessa_id"].dropna():
                match = re.search(r"R(\d+)-", str(remessa))
                if match:
                    numeros.append(int(match.group(1)))

        proximo_numero = max(numeros) + 1 if numeros else 1

    except:
        proximo_numero = 1

    horario_br = datetime.now(ZoneInfo("America/Sao_Paulo")).strftime("%H:%M")

    return f"R{proximo_numero}-{horario_br}"
    
def br_to_float(valor):
    if valor is None:
        return 0.0

    valor = str(valor).strip()
    valor = valor.replace("R$", "").strip()
    valor = valor.replace(".", "").replace(",", ".")

    try:
        return float(valor)
    except:
        return 0.0

def format_data_br(data):
    if pd.isna(data) or data is None:
        return ""
    if isinstance(data, str):
        return data
    return data.strftime("%d/%m/%Y")


def calcular_data_previsao(data_vencimento):
    if data_vencimento.weekday() == 5:  # sábado
        return data_vencimento + timedelta(days=2)

    if data_vencimento.weekday() == 6:  # domingo
        return data_vencimento + timedelta(days=1)

    return data_vencimento


def primeiro_dia_mes_atual():
    hoje = date.today()
    return date(hoje.year, hoje.month, 1)


def extrair_texto_pdf(arquivo_pdf):
    texto = ""

    with fitz.open(stream=arquivo_pdf.read(), filetype="pdf") as doc:
        for page in doc:
            texto += page.get_text("text") + "\n"

    return texto


def extrair_boleto(arquivo_pdf, tipo_operacao):
    nome_arquivo = arquivo_pdf.name
    texto = extrair_texto_pdf(arquivo_pdf)

    # Código de barras / linha digitável
    codigo_barras = ""
    match_codigo = re.search(
        r"(341\d{2}\.\d{5}\s+\d{5}\.\d{6}\s+\d{5}\.\d{6}\s+\d\s+\d{14})",
        texto
    )
    if match_codigo:
        codigo_barras = match_codigo.group(1).strip()

    # Datas
    datas = re.findall(r"\b\d{2}/\d{2}/\d{4}\b", texto)

    data_documento = None
    vencimento = None

    if len(datas) >= 2:
        vencimento = datetime.strptime(datas[0], "%d/%m/%Y").date()
        data_documento = datetime.strptime(datas[1], "%d/%m/%Y").date()

    # Valor do documento
    valor_documento = 0.0
    match_valor = re.search(r"R\$\s*([\d\.]+,\d{2})", texto)
    if match_valor:
        valor_documento = br_to_float(match_valor.group(1))
    else:
        valores = re.findall(r"\b\d{1,3}(?:\.\d{3})*,\d{2}\b", texto)
        if valores:
            valor_documento = br_to_float(valores[0])

    # Pagador
    pagador = ""
    
    match_pagador = re.search(
        r"Pagador:\s*([A-Za-zÀ-ÿ\s]+?)\s+CPF:",
        texto,
        re.IGNORECASE
    )
    
    if match_pagador:
        pagador = match_pagador.group(1).strip()
    else:
        # fallback: pega a linha anterior ao CPF
        linhas = [l.strip() for l in texto.splitlines() if l.strip()]
    
        for i, linha in enumerate(linhas):
            if "CPF:" in linha and i > 0:
                possivel_nome = linhas[i - 1].replace("Pagador:", "").strip()
    
                if possivel_nome and not possivel_nome.upper().startswith("ENDERE"):
                    pagador = possivel_nome
                    break    

    data_registro = primeiro_dia_mes_atual()
    
    data_atual_importacao = data_hoje_br_date()

    vencimento = data_atual_importacao
    data_previsao = data_atual_importacao
    data_pagamento = None
    
    config = buscar_config_padrao(tipo_operacao)

    observacoes = config["texto_observacao"] + pagador

    print("PAGADOR EXTRAÍDO:", pagador)

    return {
        "nome_arquivo": nome_arquivo,
        "data_documento": data_documento,
        "vencimento": vencimento,
        "valor_documento": valor_documento,
        "codigo_barras": codigo_barras,
        "pagador": pagador,
        "tipo_operacao": tipo_operacao,
        "data_registro": data_registro,
        "data_previsao": data_previsao,
        "data_pagamento": data_pagamento,
        "observacoes": observacoes,
        "status": "extraido"
    }


def salvar_no_supabase(df, tamanho_lote=100):
    dados = []
    remessa_id = gerar_remessa_id()

    for _, row in df.iterrows():
        dados.append({
            "remessa_id": remessa_id,
            "tipo_operacao": row["tipo_operacao"],
            "nome_arquivo": row["nome_arquivo"],
            "data_documento": row["data_documento"].isoformat() if row["data_documento"] else None,
            "vencimento": row["vencimento"].isoformat() if row["vencimento"] else None,
            "valor_documento": float(row["valor_documento"] or 0),
            "codigo_barras": row["codigo_barras"],
            "pagador": row["pagador"],
            "data_registro": row["data_registro"].isoformat() if row["data_registro"] else None,
            "data_previsao": row["data_previsao"].isoformat() if row["data_previsao"] else None,
            "data_pagamento": row["data_pagamento"].isoformat() if row["data_pagamento"] else None,
            "observacoes": row["observacoes"],
            "status": "salvo"
        })

    try:
        for i in range(0, len(dados), tamanho_lote):
            lote = dados[i:i + tamanho_lote]
            supabase.table("boletos_extraidos").insert(lote).execute()

        return True, f"{len(dados)} boleto(s) salvo(s) com sucesso na remessa {remessa_id}.", remessa_id

    except Exception as e:
        return False, f"Erro ao salvar. Possível boleto duplicado pelo código de barras. Detalhe: {e}", None

def normalizar_texto(txt):
    if txt is None:
        return ""

    txt = str(txt).strip().upper()
    txt = txt.replace("\n", " ")
    txt = txt.replace("*", "")
    txt = " ".join(txt.split())

    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in txt if not unicodedata.combining(c))

    return txt


def localizar_colunas_por_cabecalho(ws, linha_cabecalho=5):
    mapa = {}

    for col in range(1, ws.max_column + 1):
        valor = ws.cell(linha_cabecalho, col).value

        if valor:
            mapa[normalizar_texto(valor)] = col

    return mapa


def escrever_por_coluna(ws, linha, mapa_colunas, nome_coluna, valor):
    chave = normalizar_texto(nome_coluna)

    if chave not in mapa_colunas:
        raise Exception(f"Coluna não encontrada na planilha: {nome_coluna}")

    ws.cell(linha, mapa_colunas[chave]).value = valor

def copiar_estilo_linha(ws, linha_modelo, linha_destino):
    ws.row_dimensions[linha_destino].height = ws.row_dimensions[linha_modelo].height

    for col in range(1, ws.max_column + 1):
        origem = ws.cell(linha_modelo, col)
        destino = ws.cell(linha_destino, col)

        if origem.has_style:
            destino._style = copy(origem._style)

        destino.font = copy(origem.font)
        destino.fill = copy(origem.fill)
        destino.border = copy(origem.border)
        destino.alignment = copy(origem.alignment)
        destino.number_format = origem.number_format
        destino.protection = copy(origem.protection)

def data_hoje_br():
    return datetime.now(ZoneInfo("America/Sao_Paulo")).strftime("%d/%m/%Y")    
        
def gerar_excel_omie(df, tipo_operacao):
    wb = load_workbook(MODELO_EXCEL)
    ws = wb["Omie_Contas_Pagar"]
    
    config = buscar_config_padrao(tipo_operacao)
    
    data_exportacao = data_hoje_br()

    linha_cabecalho = 5
    linha_inicial = 6

    mapa_colunas = localizar_colunas_por_cabecalho(ws, linha_cabecalho)

    for idx, row in df.iterrows():
        linha = linha_inicial + idx
        copiar_estilo_linha(ws, linha_inicial, linha)

        escrever_por_coluna(
            ws, linha, mapa_colunas,
            "Fornecedor * (Razão Social, Nome Fantasia, CNPJ ou CPF)",
            config["fornecedor"]
        )

        escrever_por_coluna(
            ws, linha, mapa_colunas,
            "Categoria *",
            config["categoria"]
        )

        escrever_por_coluna(
            ws, linha, mapa_colunas,
            "Conta Corrente *",
            config["conta_corrente"]
        )

        escrever_por_coluna(
            ws, linha, mapa_colunas,
            "Data de Emissão",
            data_para_excel_br(row["data_documento"])
        )

        escrever_por_coluna(
            ws, linha, mapa_colunas,
            "Data de Registro *",
            data_para_excel_br(row["data_registro"])
        )

        escrever_por_coluna(
            ws, linha, mapa_colunas,
            "Data de Vencimento",
            data_para_excel_br(row["vencimento"])
        )

        escrever_por_coluna(
            ws, linha, mapa_colunas,
            "Data de Previsão",
            data_para_excel_br(row["data_previsao"])
        )

        escrever_por_coluna(
            ws, linha, mapa_colunas,
            "Data do Pagamento",
            ""
        )

        escrever_por_coluna(
            ws, linha, mapa_colunas,
            "Valor do Pagamento",
            ""
        )

        escrever_por_coluna(
            ws, linha, mapa_colunas,
            "Valor da Conta",
            float(row["valor_documento"] or 0)
        )

        escrever_por_coluna(
            ws, linha, mapa_colunas,
            "Observações",
            row["observacoes"]
        )

        escrever_por_coluna(
            ws, linha, mapa_colunas,
            "Tipo de Documento",
            config["tipo_documento"]
        )

        escrever_por_coluna(
            ws, linha, mapa_colunas,
            "Forma de Pagamento",
            config["forma_pagamento"]
        )

        escrever_por_coluna(
            ws, linha, mapa_colunas,
            "Código de Barras do Boleto",
            row["codigo_barras"]
        )

        escrever_por_coluna(
            ws, linha, mapa_colunas,
            "Departamento (100%)",
            config["departamento"]
        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output

def data_para_excel_br(valor):
    if pd.isna(valor) or valor is None:
        return ""

    data = pd.to_datetime(valor, errors="coerce")

    if pd.isna(data):
        return ""

    return data.strftime("%d/%m/%Y")

def preparar_df_visual(df):
    df_view = df.copy()

    colunas_data = [
        "data_documento",
        "vencimento",
        "data_registro",
        "data_previsao",
        "data_pagamento",
        "created_at"
    ]

    for col in colunas_data:
        if col in df_view.columns:
            df_view[col] = pd.to_datetime(df_view[col], errors="coerce").dt.strftime("%d/%m/%Y")

    nomes_colunas = {
        "id": "ID",
        "remessa_id": "Remessa",
        "nome_arquivo": "Nome Arquivo",
        "data_documento": "Data Documento",
        "vencimento": "Vencimento",
        "valor_documento": "Valor Documento",
        "codigo_barras": "Código Barras",
        "pagador": "Pagador",
        "data_registro": "Data Registro",
        "data_previsao": "Data Previsão",
        "data_pagamento": "Data Pagamento",
        "observacoes": "Observações",
        "status": "Status",
        "created_at": "Criado Em"
    }

    df_view = df_view.rename(columns=nomes_colunas)

    return df_view
    
# =============================
# INTERFACE / MENU
# =============================

st.markdown("""
<style>
    .main-title {
        font-size: 32px;
        font-weight: 700;
        color: #1f2937;
    }

    .subtitle {
        font-size: 15px;
        color: #6b7280;
        margin-bottom: 20px;
    }

    .card {
        background-color: white;
        padding: 22px;
        border-radius: 14px;
        border: 1px solid #e5e7eb;
        box-shadow: 0px 2px 8px rgba(0,0,0,0.05);
        text-align: center;
    }

    .card-title {
        font-size: 14px;
        color: #6b7280;
    }

    .card-value {
        font-size: 28px;
        font-weight: 700;
        color: #111827;
    }

    section[data-testid="stSidebar"] {
        background-color: #ffffff;
        border-right: 1px solid #e5e7eb;
    }
    
    section[data-testid="stSidebar"] * {
        color: #111827;
        
    }
</style>
""", unsafe_allow_html=True)


with st.sidebar:
    st.image("assets/logo.png", width=180)
    st.markdown("")
    st.caption("Conciliação GooRoo")

    pagina = st.radio(
        "Menu",
        [
            "Dashboard",
            "Importar Boletos",
            "Histórico de Remessas",
            "Gerar Planilha Omie",
            "Configurações"
        ]
    )


# =============================
# DASHBOARD
# =============================
if pagina == "Dashboard":

    st.markdown('<div class="main-title">📊 Dashboard</div>', unsafe_allow_html=True)
    st.markdown('<div class="subtitle">Visão geral dos boletos e filtro por dia de importação.</div>', unsafe_allow_html=True)

    try:
        res = supabase.table("boletos_extraidos").select("*").execute()
        df_dash = pd.DataFrame(res.data)

        if df_dash.empty:
            st.info("Nenhum boleto salvo ainda.")
        else:
            df_dash["created_at"] = pd.to_datetime(df_dash["created_at"], errors="coerce")
            df_dash["data_importacao"] = df_dash["created_at"].dt.date

            # =============================
            # VISÃO TOTAL
            # =============================
            st.markdown("### 📌 Visão Geral Total")

            total_boletos_geral = len(df_dash)
            valor_total_geral = df_dash["valor_documento"].sum() if "valor_documento" in df_dash.columns else 0
            total_remessas_geral = (
                df_dash["remessa_id"].nunique()
                if "remessa_id" in df_dash.columns
                else 0
            )

            col1, col2, col3 = st.columns(3)

            with col1:
                st.markdown(f"""
                <div class="card">
                    <div class="card-title">Total de Boletos</div>
                    <div class="card-value">{total_boletos_geral}</div>
                </div>
                """, unsafe_allow_html=True)

            with col2:
                valor_formatado_geral = f"R$ {valor_total_geral:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

                st.markdown(f"""
                <div class="card">
                    <div class="card-title">Valor Total Geral</div>
                    <div class="card-value">{valor_formatado_geral}</div>
                </div>
                """, unsafe_allow_html=True)

            with col3:
                st.markdown(f"""
                <div class="card">
                    <div class="card-title">Total de Remessas</div>
                    <div class="card-value">{total_remessas_geral}</div>
                </div>
                """, unsafe_allow_html=True)

            st.markdown("---")

            # =============================
            # FILTRO POR DIA
            # =============================
            st.markdown("### 📅 Filtro por Dia")

            data_filtro = st.date_input(
                "Selecione o dia de importação",
                value=date.today(),
                format="DD/MM/YYYY"
            )

            df_filtrado = df_dash[df_dash["data_importacao"] == data_filtro].copy()

            total_boletos_dia = len(df_filtrado)
            valor_total_dia = df_filtrado["valor_documento"].sum() if not df_filtrado.empty else 0

            remessas_dia = (
                df_filtrado["remessa_id"].nunique()
                if "remessa_id" in df_filtrado.columns and not df_filtrado.empty
                else 0
            )

            col4, col5, col6 = st.columns(3)

            with col4:
                st.markdown(f"""
                <div class="card">
                    <div class="card-title">Boletos no Dia</div>
                    <div class="card-value">{total_boletos_dia}</div>
                </div>
                """, unsafe_allow_html=True)

            with col5:
                valor_formatado_dia = f"R$ {valor_total_dia:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

                st.markdown(f"""
                <div class="card">
                    <div class="card-title">Valor Total no Dia</div>
                    <div class="card-value">{valor_formatado_dia}</div>
                </div>
                """, unsafe_allow_html=True)

            with col6:
                st.markdown(f"""
                <div class="card">
                    <div class="card-title">Remessas no Dia</div>
                    <div class="card-value">{remessas_dia}</div>
                </div>
                """, unsafe_allow_html=True)

            st.markdown("### Boletos importados no dia")

            if df_filtrado.empty:
                st.warning("Nenhum boleto encontrado para esta data.")
            else:
                colunas_exibir = [
                    "remessa_id",
                    "nome_arquivo",
                    "data_documento",
                    "vencimento",
                    "valor_documento",
                    "codigo_barras",
                    "pagador",
                    "data_previsao",
                    "data_pagamento",
                    "created_at"
                ]

                colunas_exibir = [c for c in colunas_exibir if c in df_filtrado.columns]

                st.dataframe(
                    preparar_df_visual(df_filtrado[colunas_exibir]),
                    use_container_width=True
                )

                csv = df_filtrado[colunas_exibir].to_csv(
                    index=False,
                    sep=";",
                    encoding="utf-8-sig"
                ).encode("utf-8-sig")

                st.download_button(
                    label="📥 Baixar extração do dia em CSV",
                    data=csv,
                    file_name=f"boletos_importados_{data_filtro.strftime('%d_%m_%Y')}.csv",
                    mime="text/csv",
                    use_container_width=True
                )

    except Exception as e:
        st.error(f"Erro ao carregar dashboard: {e}")


# =============================
# IMPORTAR BOLETOS
# =============================
elif pagina == "Importar Boletos":

    st.markdown('<div class="main-title">📄 Importar Boletos</div>', unsafe_allow_html=True)
    st.markdown('<div class="subtitle">Selecione vários boletos em PDF para extração automática.</div>', unsafe_allow_html=True)

    if "upload_key" not in st.session_state:
        st.session_state["upload_key"] = 0

    tipo_operacao = st.radio(
        "Tipo de operação",
        ["SEGUROS", "SERASA"],
        horizontal=True
    )

    arquivos = st.file_uploader(
        "Selecione os boletos em PDF",
        type=["pdf"],
        accept_multiple_files=True,
        key=f"upload_boletos_{st.session_state['upload_key']}"
    )

    if arquivos:
        if st.button("❌ Cancelar / Limpar seleção", use_container_width=True):
            st.session_state["upload_key"] += 1
    
            for chave in [
                "df_boletos",
                "df_boletos_editado",
                "arquivos_anteriores"
            ]:
                if chave in st.session_state:
                    del st.session_state[chave]
    
            st.rerun()

    arquivos_atuais = [arquivo.name for arquivo in arquivos] if arquivos else []

    if st.session_state.get("arquivos_anteriores") != arquivos_atuais:
        st.session_state["arquivos_anteriores"] = arquivos_atuais
    
        if "df_boletos" in st.session_state:
            del st.session_state["df_boletos"]
    
        if "df_boletos_editado" in st.session_state:
            del st.session_state["df_boletos_editado"]

    if arquivos:
        st.success(f"{len(arquivos)} arquivo(s) selecionado(s).")

        if st.button("🔎 Extrair informações dos boletos", use_container_width=True):
            registros = []
            progresso = st.progress(0)

            for i, arquivo in enumerate(arquivos):
                try:
                    dados = extrair_boleto(arquivo, tipo_operacao)
                    registros.append(dados)
                except Exception as e:
                    st.error(f"Erro ao processar {arquivo.name}: {e}")

                progresso.progress((i + 1) / len(arquivos))

            st.session_state["df_boletos"] = pd.DataFrame(registros)

    if "df_boletos" in st.session_state:
        st.markdown("### ✅ Conferência dos dados extraídos")

        df_visual_importacao = preparar_df_visual(
            st.session_state["df_boletos"]
        )
        
        df_editado_visual = st.data_editor(
            df_visual_importacao,
            use_container_width=True,
            num_rows="dynamic"
        )

        df_editado = st.session_state["df_boletos"]

        if not df_editado.empty and "valor_documento" in df_editado.columns:
            valor_total_importacao = pd.to_numeric(
                df_editado["valor_documento"],
                errors="coerce"
            ).fillna(0).sum()
        else:
            valor_total_importacao = 0
        
        valor_total_formatado = (
            f"R$ {valor_total_importacao:,.2f}"
            .replace(",", "X")
            .replace(".", ",")
            .replace("X", ".")
        )
        
        st.success(f"💰 Valor total dos boletos extraídos: **{valor_total_formatado}**")

        if st.button("💾 Salvar Extração", use_container_width=True):
            with st.spinner("Salvando boletos no Supabase em lotes..."):
                ok, msg, remessa_id = salvar_no_supabase(df_editado, tamanho_lote=100)
        
            if ok:
                st.session_state["ultima_remessa_id"] = remessa_id
                st.success(msg)
            else:
                st.error(msg)      

# =============================
# HISTÓRICO DE REMESSAS
# =============================
elif pagina == "Histórico de Remessas":

    st.markdown('<div class="main-title">🧾 Histórico de Remessas</div>', unsafe_allow_html=True)
    st.markdown('<div class="subtitle">Resumo das remessas importadas no sistema.</div>', unsafe_allow_html=True)

    try:
        res = (
            supabase.table("boletos_extraidos")
            .select("*")
            .order("created_at", desc=True)
            .execute()
        )

        df = pd.DataFrame(res.data)

        if df.empty:
            st.info("Nenhuma remessa encontrada.")
        else:
            df["created_at"] = pd.to_datetime(df["created_at"], errors="coerce")
            df["valor_documento"] = pd.to_numeric(df["valor_documento"], errors="coerce").fillna(0)

            resumo = (
                df.groupby(["remessa_id", "tipo_operacao"], dropna=False)
                .agg(
                    Data_Importacao=("created_at", "max"),
                    Quantidade=("id", "count"),
                    Valor_Total=("valor_documento", "sum")
                )
                .reset_index()
                .sort_values("Data_Importacao", ascending=False)
            )

            resumo["Data_Importacao"] = (
                resumo["Data_Importacao"]
                .dt.tz_convert("America/Sao_Paulo")
                .dt.strftime("%d/%m/%Y %H:%M")
            )
            
            resumo["Valor_Total"] = resumo["Valor_Total"].apply(
                lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            )

            resumo = resumo.rename(columns={
                "remessa_id": "Remessa",
                "tipo_operacao": "Tipo Operação",
                "Data_Importacao": "Data Importação",
                "Valor_Total": "Valor Total"
            })

            st.dataframe(resumo, use_container_width=True)

            st.markdown("### 📥 Extrair relatório por remessa")

            remessa_escolhida_hist = st.selectbox(
                "Selecione a remessa para extração",
                resumo["Remessa"].tolist()
            )
            
            df_remessa = df[df["remessa_id"] == remessa_escolhida_hist].copy()
            
            st.write(f"Boletos encontrados: **{len(df_remessa)}**")
            
            csv_remessa = df_remessa.to_csv(
                index=False,
                sep=";",
                encoding="utf-8-sig"
            ).encode("utf-8-sig")
            
            st.download_button(
                label="📥 Baixar relatório da remessa em CSV",
                data=csv_remessa,
                file_name=f"relatorio_remessa_{remessa_escolhida_hist}.csv",
                mime="text/csv",
                use_container_width=True
            )

    except Exception as e:
        st.error(f"Erro ao carregar histórico de remessas: {e}")

# =============================
# GERAR PLANILHA OMIE
# =============================
elif pagina == "Gerar Planilha Omie":

    st.markdown('<div class="main-title">📥 Gerar Planilha Omie</div>', unsafe_allow_html=True)
    st.markdown('<div class="subtitle">Gere a planilha Omie por remessa importada.</div>', unsafe_allow_html=True)

    try:
        res_remessas = (
            supabase.table("boletos_extraidos")
            .select("remessa_id, created_at")
            .not_.is_("remessa_id", "null")
            .order("created_at", desc=True)
            .execute()
        )

        df_remessas = pd.DataFrame(res_remessas.data)

        if df_remessas.empty:
            st.info("Nenhuma remessa disponível para gerar planilha.")
        else:
            df_remessas["created_at"] = pd.to_datetime(df_remessas["created_at"], errors="coerce")

            resumo_remessas = (
                df_remessas
                .dropna(subset=["remessa_id"])
                .groupby("remessa_id")
                .agg(data_importacao=("created_at", "max"))
                .reset_index()
                .sort_values("data_importacao", ascending=False)
            )

            opcoes = resumo_remessas["remessa_id"].tolist()

            remessa_padrao = st.session_state.get("ultima_remessa_id")

            index_padrao = 0
            if remessa_padrao in opcoes:
                index_padrao = opcoes.index(remessa_padrao)

            remessa_escolhida = st.selectbox(
                "Selecione a remessa",
                opcoes,
                index=index_padrao
            )

            res = (
                supabase.table("boletos_extraidos")
                .select("*")
                .eq("remessa_id", remessa_escolhida)
                .order("created_at", desc=False)
                .execute()
            )

            df = pd.DataFrame(res.data)

            if df.empty:
                st.info("Nenhum boleto encontrado para esta remessa.")
            else:
                total = len(df)
                valor_total = df["valor_documento"].sum()

                col1, col2 = st.columns(2)

                with col1:
                    st.metric("Boletos da Remessa", total)

                with col2:
                    st.metric(
                        "Valor Total",
                        f"R$ {valor_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                    )

                st.dataframe(preparar_df_visual(df), use_container_width=True)

                tipo_operacao_remessa = df["tipo_operacao"].iloc[0]
                excel = gerar_excel_omie(df, tipo_operacao_remessa)

                st.download_button(
                    label="📥 Baixar planilha Omie desta remessa",
                    data=excel,
                    file_name=f"Omie_Contas_Pagar_Remessa_{remessa_escolhida}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

    except Exception as e:
        st.error(f"Erro ao gerar planilha: {e}")


# =============================
# CONFIGURAÇÕES
# =============================
elif pagina == "Configurações":

    st.markdown('<div class="main-title">⚙️ Configurações</div>', unsafe_allow_html=True)
    st.info("Aqui depois vamos editar os padrões: fornecedor, categoria, conta corrente e departamento.")
